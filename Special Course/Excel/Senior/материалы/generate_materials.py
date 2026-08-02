# -*- coding: utf-8 -*-
"""
Генератор учебных файлов для курса «Excel для работы».

Запуск:  python generate_materials.py
Требует: openpyxl  (pip install openpyxl)

Создаёт в текущей папке правдоподобную «рабочую» выгрузку торговой компании «ТехноТрейд»:
  - Продажи.xlsx              заказы (300 строк) — основная таблица курса
  - Прайс.xlsx                справочник товар→категория→цена (для ВПР/ПРОСМОТРX)
  - Планы.xlsx                план выручки по менеджерам (для план-факт анализа)
  - Сотрудники.xlsx           «грязные» данные кадров (ФИО, даты-текст, телефоны)
  - Выгрузка_грязная.csv      грязный CSV для чистки в Power Query
  - Данные_для_дашборда.xlsx  заказы за год (1200 строк) — для итогового дашборда
  - Выгрузки_по_месяцам/      3 помесячных CSV (для сценария Power Query «собери из папки»)

Данные детерминированы (random.seed=42). Столбцы A–H в «Продажи»/«Данные_для_дашборда»
совпадают с прежней версией; реалистичные поля (№ заказа, клиент, канал, отгрузка,
статус оплаты) добавлены СПРАВА (столбцы I–N), чтобы прежние уроки не «поехали».
Все данные вымышленные.
"""
import csv
import os
import random
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

random.seed(42)

# ---------- Справочники ----------
MANAGERS = ["Иванова А.", "Петров С.", "Сидорова М.", "Кузнецов Д.", "Смирнова О.", "Волков И."]
CITIES = ["Москва", "Санкт-Петербург", "Казань", "Новосибирск", "Екатеринбург", "Краснодар"]
CHANNELS = ["Онлайн-магазин", "Розничный магазин", "Телефон", "Менеджер"]
CLIENTS = ["ООО Ромашка", "ИП Соколов", "ООО ТехноПлюс", "АО Вектор", "ООО Мир Офиса",
           "ИП Гаврилов", "ООО Стройсервис", "АО Прогресс", "ООО Аметист", "Розничный клиент"]
PAY_STATUSES = (["Оплачен"] * 6) + (["Ожидает оплаты"] * 3) + (["Просрочен"] * 1)

# Товар -> (категория, цена)
PRODUCTS = {
    "Ноутбук Aspire 15": ("Компьютеры", 54990),
    "Ноутбук ProBook 14": ("Компьютеры", 72990),
    "Монитор 27\" IPS": ("Периферия", 18990),
    "Клавиатура механическая": ("Периферия", 4990),
    "Мышь беспроводная": ("Периферия", 1490),
    "Веб-камера Full HD": ("Периферия", 3990),
    "Наушники с шумоподавлением": ("Аудио", 12990),
    "Колонка Bluetooth": ("Аудио", 5990),
    "Микрофон USB": ("Аудио", 8990),
    "SSD 1 ТБ": ("Комплектующие", 8490),
    "Оперативная память 16 ГБ": ("Комплектующие", 4290),
    "Видеокарта RTX": ("Комплектующие", 64990),
    "Роутер Wi-Fi 6": ("Сеть", 6990),
    "Сетевой фильтр": ("Аксессуары", 990),
    "Коврик для мыши XL": ("Аксессуары", 790),
    "Сумка для ноутбука": ("Аксессуары", 2490),
}
PRODUCT_NAMES = list(PRODUCTS.keys())

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def autofit(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 3, 42)


# ---------- Продажи / Данные для дашборда ----------
def make_sales(filename, n, start, days, first_order_no):
    """A–H — как раньше (Сумма пустая). I–N — реалистичные поля."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Продажи"
    headers = ["Дата заказа", "Менеджер", "Город", "Категория", "Товар",
               "Количество", "Цена", "Сумма",
               "№ заказа", "Клиент", "Тип клиента", "Канал продаж",
               "Дата отгрузки", "Статус оплаты"]
    ws.append(headers)
    for i in range(n):
        # --- базовые столбцы A–H: тот же порядок случайных вызовов, что и раньше ---
        d = start + timedelta(days=random.randint(0, days))
        product = random.choice(PRODUCT_NAMES)
        category, price = PRODUCTS[product]
        qty = random.randint(1, 8)
        manager = random.choice(MANAGERS)
        city = random.choice(CITIES)
        # --- новые столбцы I–N (доп. случайные вызовы ПОСЛЕ базовых) ---
        client = random.choice(CLIENTS)
        client_type = "Опт" if qty >= 5 else "Розница"
        channel = random.choice(CHANNELS)
        ship = d + timedelta(days=random.randint(1, 7))
        pay = random.choice(PAY_STATUSES)
        order_no = first_order_no + i
        ws.append([d, manager, city, category, product, qty, price, None,
                   order_no, client, client_type, channel, ship, pay])
    for r in range(2, n + 2):
        ws.cell(row=r, column=1).number_format = "DD.MM.YYYY"
        ws.cell(row=r, column=7).number_format = "# ##0 ₽"
        ws.cell(row=r, column=13).number_format = "DD.MM.YYYY"
    style_header(ws, len(headers))
    autofit(ws)
    wb.save(filename)
    print(f"  сохранён {filename} ({n} строк, 14 столбцов)")


# ---------- Прайс ----------
def make_price(filename="Прайс.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Прайс"
    ws.append(["Товар", "Категория", "Цена"])
    for name, (cat, price) in PRODUCTS.items():
        ws.append([name, cat, price])
    for r in range(2, len(PRODUCTS) + 2):
        ws.cell(row=r, column=3).number_format = "# ##0 ₽"
    style_header(ws, 3)
    autofit(ws)
    wb.save(filename)
    print(f"  сохранён {filename} ({len(PRODUCTS)} товаров)")


# ---------- Планы (план-факт по менеджерам) ----------
def make_plans(filename="Планы.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Планы"
    ws.append(["Менеджер", "План выручки, ₽"])
    plans = {  # округлённые «плановые» цифры на период
        "Иванова А.": 2_500_000, "Петров С.": 2_200_000, "Сидорова М.": 2_400_000,
        "Кузнецов Д.": 2_000_000, "Смирнова О.": 2_300_000, "Волков И.": 1_900_000,
    }
    for m in MANAGERS:
        ws.append([m, plans[m]])
    for r in range(2, len(MANAGERS) + 2):
        ws.cell(row=r, column=2).number_format = "# ##0 ₽"
    style_header(ws, 2)
    autofit(ws)
    wb.save(filename)
    print(f"  сохранён {filename} (план по {len(MANAGERS)} менеджерам)")


# ---------- Сотрудники («грязные» данные) ----------
def make_employees(filename="Сотрудники.xlsx", n=40):
    last_m = ["Иванов", "Петров", "Сидоров", "Кузнецов", "Смирнов", "Волков", "Морозов", "Новиков"]
    last_f = ["Иванова", "Петрова", "Сидорова", "Кузнецова", "Смирнова", "Волкова", "Морозова", "Новикова"]
    first_m = ["Александр", "Сергей", "Дмитрий", "Иван", "Пётр", "Николай"]
    first_f = ["Анна", "Мария", "Ольга", "Елена", "Наталья", "Ирина"]
    patr_m = ["Александрович", "Сергеевич", "Дмитриевич", "Иванович", "Петрович"]
    patr_f = ["Александровна", "Сергеевна", "Дмитриевна", "Ивановна", "Петровна"]
    depts = ["Продажи", "Маркетинг", "Логистика", "Бухгалтерия", "IT"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"
    ws.append(["ФИО", "Отдел", "Дата приёма", "Телефон", "Email"])
    for _ in range(n):
        if random.random() < 0.5:
            ln, fn, pt = random.choice(last_m), random.choice(first_m), random.choice(patr_m)
        else:
            ln, fn, pt = random.choice(last_f), random.choice(first_f), random.choice(patr_f)
        fio = f"{ln} {fn} {pt}"
        if random.random() < 0.3:
            fio = "  " + fio + " "
        if random.random() < 0.2:
            fio = fio.replace(" ", "  ", 1)
        d = date(random.randint(2015, 2024), random.randint(1, 12), random.randint(1, 28))
        digits = f"9{random.randint(10,99)}{random.randint(1000000,9999999)}"
        fmt = random.choice([
            f"+7 ({digits[0:3]}) {digits[3:6]}-{digits[6:8]}-{digits[8:10]}",
            f"8{digits}",
            f"+7{digits}",
            f"7 {digits[0:3]} {digits[3:]}",
        ])
        email = f"{fn.lower()}.{ln.lower()}@company.ru"
        ws.append([fio, random.choice(depts), d.strftime("%d.%m.%Y"), fmt, email])
    style_header(ws, 5)
    autofit(ws)
    wb.save(filename)
    print(f"  сохранён {filename} ({n} сотрудников, 'грязные' данные)")


# ---------- Грязный CSV (для очистки в Power Query) ----------
def make_dirty_csv(filename="Выгрузка_грязная.csv", n=120):
    city_variants = {
        "Москва": ["Москва", "москва", " Москва", "МОСКВА", "Москва "],
        "Санкт-Петербург": ["Санкт-Петербург", "СПб", "санкт-петербург", " Санкт-Петербург"],
        "Казань": ["Казань", "казань", "Казань "],
    }
    base_cities = list(city_variants.keys())
    rows = [["Дата", "Город", "Товар", "Количество", "Сумма"]]
    for _ in range(n):
        if random.random() < 0.08:
            rows.append(["", "", "", "", ""])
            continue
        d = date(2025, random.randint(1, 6), random.randint(1, 28))
        city = random.choice(city_variants[random.choice(base_cities)])
        product = random.choice(PRODUCT_NAMES)
        qty = random.randint(1, 10)
        _, price = PRODUCTS[product]
        total = qty * price
        total_str = f" {total:,} ".replace(",", " ") if random.random() < 0.5 else str(total)
        rows.append([d.strftime("%d.%m.%Y"), city, "  " + product if random.random() < 0.3 else product,
                     str(qty), total_str])
    with open(filename, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f, delimiter=";").writerows(rows)
    print(f"  сохранён {filename} ({n} строк, 'грязный')")


# ---------- Помесячные выгрузки (Power Query: собрать из папки) ----------
def make_monthly_exports(folder="Выгрузки_по_месяцам"):
    os.makedirs(folder, exist_ok=True)
    months = [("2025-01", date(2025, 1, 1), 31),
              ("2025-02", date(2025, 2, 1), 28),
              ("2025-03", date(2025, 3, 1), 31)]
    for name, start, days in months:
        rows = [["Дата", "Менеджер", "Город", "Товар", "Количество", "Цена"]]
        for _ in range(random.randint(35, 50)):
            d = start + timedelta(days=random.randint(0, days - 1))
            product = random.choice(PRODUCT_NAMES)
            _, price = PRODUCTS[product]
            rows.append([d.strftime("%d.%m.%Y"), random.choice(MANAGERS), random.choice(CITIES),
                         product, random.randint(1, 8), price])
        path = os.path.join(folder, f"{name}.csv")
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            csv.writer(f, delimiter=";").writerows(rows)
    print(f"  сохранены помесячные выгрузки в папке {folder}/ (3 файла)")


if __name__ == "__main__":
    print("Генерация учебных файлов…")
    make_sales("Продажи.xlsx", n=300, start=date(2025, 1, 1), days=180, first_order_no=100001)
    make_price("Прайс.xlsx")
    make_plans("Планы.xlsx")
    make_employees("Сотрудники.xlsx", n=40)
    make_dirty_csv("Выгрузка_грязная.csv", n=120)
    make_sales("Данные_для_дашборда.xlsx", n=1200, start=date(2025, 1, 1), days=364, first_order_no=200001)
    make_monthly_exports("Выгрузки_по_месяцам")
    print("Готово. Файлы лежат рядом со скриптом.")
